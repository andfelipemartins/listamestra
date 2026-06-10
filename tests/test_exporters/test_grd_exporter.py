"""
tests/test_exporters/test_grd_exporter.py

Testes dos exportadores de GRD (Excel e PDF) — geram bytes válidos.
"""

import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from openpyxl import load_workbook
from core.exporters.grd_exporter import exportar_grd_excel, exportar_grd_pdf


def _dados():
    return {
        "cabecalho": {
            "numero_grd": "GRD-001/2026", "data_envio": "2026-06-07",
            "destinatario": "METRÔ-SP", "ac": "Eng. Fulano", "obra": "Linha 15",
            "trecho": "25 — Ragueb Chohfi", "status": "recebida",
            "observacoes": "Remessa inicial",
            "emitido_por": "Maria Emissora", "recebido_por": "João Recebedor",
            "recebido_cargo": "Engenheiro", "recebido_em": "2026-06-10",
            "declaracao_recebimento": "Recebi os documentos.",
        },
        "itens": [
            {
                "codigo": "DE-15.25.00.00-6A1-1001", "titulo": "Planta Geral",
                "label_revisao": "0", "versao": 1, "situacao": "APROVADO",
                "qtd_a0": 1, "qtd_a1": 2, "qtd_a2": 0, "qtd_a3": 0, "qtd_a4": 4, "qtd_digital": 1,
            },
            {
                "codigo": "DE-15.25.00.00-6A1-1002", "titulo": "Corte AA",
                "label_revisao": "A", "versao": 2, "situacao": "APROVADO",
                "qtd_a0": 0, "qtd_a1": 1, "qtd_a2": 0, "qtd_a3": 0, "qtd_a4": 0, "qtd_digital": 2,
            },
        ],
    }


class TestExcel:
    def test_retorna_bytes_nao_vazios(self):
        out = exportar_grd_excel(_dados())
        assert isinstance(out, bytes) and len(out) > 0

    def test_xlsx_abre_e_contem_numero_e_itens(self):
        out = exportar_grd_excel(_dados())
        wb = load_workbook(io.BytesIO(out))
        ws = wb.active
        textos = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        assert any("GRD-001/2026" in t for t in textos)
        assert any("DE-15.25.00.00-6A1-1001" in t for t in textos)
        assert any("DE-15.25.00.00-6A1-1002" in t for t in textos)

    def test_xlsx_contem_campos_de_recebimento(self):
        out = exportar_grd_excel(_dados())
        ws = load_workbook(io.BytesIO(out)).active
        textos = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        assert "João Recebedor" in textos
        assert "Engenheiro" in textos
        assert "Maria Emissora" in textos
        assert "Recebi os documentos." in textos  # rodapé de declaração

    def test_xlsx_anulada_mostra_motivo_no_rodape(self):
        dados = _dados()
        dados["cabecalho"]["status"] = "anulada"
        dados["cabecalho"]["motivo_anulacao"] = "documento substituído"
        ws = load_workbook(io.BytesIO(exportar_grd_excel(dados))).active
        textos = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
        assert "ANULADA" in textos and "documento substituído" in textos

    def test_sem_itens_ainda_gera(self):
        out = exportar_grd_excel({"cabecalho": {"numero_grd": "GRD-X"}, "itens": []})
        assert isinstance(out, bytes) and len(out) > 0


class TestPdf:
    def test_retorna_bytes_pdf(self):
        out = exportar_grd_pdf(_dados())
        assert isinstance(out, bytes) and out[:5] == b"%PDF-"

    def test_pdf_nao_vazio(self):
        out = exportar_grd_pdf(_dados())
        assert len(out) > 500

    def test_sem_itens_ainda_gera(self):
        out = exportar_grd_pdf({"cabecalho": {"numero_grd": "GRD-X"}, "itens": []})
        assert out[:5] == b"%PDF-"
