"""
core/exporters/grd_exporter.py

Exportação da GRD (Guia de Remessa de Documentos) em Excel e PDF.

PURPOSE: gerar arquivos .xlsx e .pdf (em bytes) de uma GRD para download.
INPUTS:  dict {cabecalho, itens} produzido por GrdService.montar_dados_exportacao.
OUTPUTS: bytes prontos para st.download_button.
DEPS:    openpyxl (Excel), reportlab (PDF).
SEE:     core/services/grd_service.py
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

_COLUNAS = [
    ("Item", 6), ("Documento Técnico", 30), ("Rev.", 6), ("Ver.", 6),
    ("Descrição", 40), ("A0", 5), ("A1", 5), ("A2", 5), ("A3", 5),
    ("A4", 5), ("Digital", 8),
]


def _linha_item(idx: int, it: dict) -> list:
    return [
        idx,
        it.get("codigo") or "",
        it.get("label_revisao") or "",
        it.get("versao") or "",
        it.get("titulo") or "",
        it.get("qtd_a0") or 0, it.get("qtd_a1") or 0, it.get("qtd_a2") or 0,
        it.get("qtd_a3") or 0, it.get("qtd_a4") or 0, it.get("qtd_digital") or 0,
    ]


def _cabecalho_pares(cab: dict, total_itens: int) -> list[tuple]:
    return [
        ("Número da GRD:", cab.get("numero_grd") or "—"),
        ("Data:", cab.get("data_envio") or "—"),
        ("Qtd. de documentos:", total_itens),
        ("Trecho:", cab.get("trecho") or "—"),
        ("Destinatário:", cab.get("destinatario") or cab.get("setor") or "—"),
        ("A/C:", cab.get("ac") or "—"),
        ("Obra:", cab.get("obra") or "—"),
        ("Status:", cab.get("status") or "—"),
    ]


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def exportar_grd_excel(dados: dict) -> bytes:
    """Gera a GRD em .xlsx (bytes). Layout de Guia de Remessa."""
    cab = dados.get("cabecalho", {}) or {}
    itens = dados.get("itens", []) or []

    wb = Workbook()
    ws = wb.active
    ws.title = "GRD"

    titulo_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    thin = Side(style="thin", color="999999")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "GUIA DE REMESSA DE DOCUMENTOS"
    ws["A1"].font = titulo_font

    linha = 3
    for label, valor in _cabecalho_pares(cab, len(itens)):
        ws.cell(row=linha, column=1, value=label).font = label_font
        ws.cell(row=linha, column=2, value=valor)
        linha += 1

    if cab.get("observacoes"):
        ws.cell(row=linha, column=1, value="Observações:").font = label_font
        ws.cell(row=linha, column=2, value=cab.get("observacoes"))
        linha += 1

    linha += 1
    head_row = linha
    for col, (nome, largura) in enumerate(_COLUNAS, start=1):
        cell = ws.cell(row=head_row, column=col, value=nome)
        cell.font = label_font
        cell.fill = head_fill
        cell.border = borda
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = largura

    for i, it in enumerate(itens, start=1):
        linha += 1
        for col, valor in enumerate(_linha_item(i, it), start=1):
            cell = ws.cell(row=linha, column=col, value=valor)
            cell.border = borda

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def exportar_grd_pdf(dados: dict) -> bytes:
    """Gera a GRD em .pdf (bytes). Mesmas informações do Excel."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet

    cab = dados.get("cabecalho", {}) or {}
    itens = dados.get("itens", []) or []
    styles = getSampleStyleSheet()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.2 * cm, bottomMargin=1.2 * cm,
        title="Guia de Remessa de Documentos",
    )

    elementos = [Paragraph("GUIA DE REMESSA DE DOCUMENTOS", styles["Title"]), Spacer(1, 8)]

    for label, valor in _cabecalho_pares(cab, len(itens)):
        elementos.append(Paragraph(f"<b>{label}</b> {valor}", styles["Normal"]))
    if cab.get("observacoes"):
        elementos.append(Paragraph(f"<b>Observações:</b> {cab.get('observacoes')}", styles["Normal"]))
    elementos.append(Spacer(1, 12))

    header = [nome for nome, _ in _COLUNAS]
    linhas = [header] + [_linha_item(i, it) for i, it in enumerate(itens, start=1)]
    tabela = Table(linhas, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (5, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabela)

    doc.build(elementos)
    buffer.seek(0)
    return buffer.read()
